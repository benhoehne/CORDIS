"""
Filtered Excel export (Task 3).
===============================

Builds an ``.xlsx`` file from the *same* :class:`ProjectFilters` the web UI
uses, so a download always matches what is on screen. Files are written into
the ``data/`` folder with a descriptive, filesystem-safe name, e.g.::

    export_erc_panels_2024_ukhd.xlsx
    export_projects_2019-2023_uhei_20260702_143005.xlsx

Naming convention
-----------------
``export_<scope>_<years>_<institution>[_<erc>]_<timestamp>.xlsx`` where each
segment is only included when the corresponding filter is set. A timestamp is
always appended to guarantee uniqueness.
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .repository import ProjectFilters, fetch_all_filtered

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = PROJECT_ROOT / "data"


def _slug(value: str) -> str:
    """Lower-case, keep alnum, collapse the rest to single underscores."""
    s = re.sub(r"[^0-9a-zA-Z]+", "_", str(value).lower()).strip("_")
    return s


def build_export_filename(f: ProjectFilters) -> str:
    """
    Derive a descriptive file name from the active filters.

    Examples
    --------
    filters(erc_only=True, panel='LS1', year_from=2024, institution='UKHD')
        -> export_erc_ls1_2024_ukhd_<ts>.xlsx
    filters(year_from=2019, year_to=2023, institution='UHEI')
        -> export_projects_2019_2023_uhei_<ts>.xlsx
    """
    parts: list[str] = ["export"]

    parts.append("erc" if f.erc_only else "projects")

    if f.panel:
        parts.append(_slug(f.panel)[:12])

    # Year range segment
    if f.year_from and f.year_to:
        parts.append(f"{f.year_from}_{f.year_to}")
    elif f.year_from:
        parts.append(f"{f.year_from}")
    elif f.year_to:
        parts.append(f"to{f.year_to}")

    if f.programme:
        parts.append(_slug(f.programme)[:12])

    if f.institution:
        parts.append(_slug(f.institution))

    if f.pi:
        parts.append(_slug(f.pi)[:16])

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    parts.append(ts)

    return "_".join(p for p in parts if p) + ".xlsx"


def export_filtered(
    filters: ProjectFilters,
    out_dir: str | Path = DATA_DIR,
    filename: str | None = None,
) -> Path:
    """
    Write all rows matching *filters* to an Excel file and return its path.

    The sheet is wrapped in a native Excel Table so the recipient can sort /
    filter immediately. Returns the :class:`Path` written.
    """
    rows = fetch_all_filtered(filters)
    df = pd.DataFrame(rows)

    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    name = filename or build_export_filename(filters)
    path = out_dir / name

    # Empty result → still write a header-only workbook so the user gets a file.
    if df.empty:
        df = pd.DataFrame(columns=["id", "acronym", "title"])

    df.to_excel(path, index=False, engine="openpyxl", sheet_name="Projects")

    # Add a native Excel table for convenience.
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb.active
    n_rows, n_cols = len(df), len(df.columns)
    if n_cols > 0:
        last_col = get_column_letter(n_cols)
        tbl = Table(displayName="HeidelbergProjects",
                    ref=f"A1:{last_col}{n_rows + 1}")
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showRowStripes=True,
        )
        ws.add_table(tbl)
        wb.save(path)

    return path
