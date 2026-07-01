"""
ERC Data Consolidation Script
==============================
Filters and merges the ERC Dashboard dump and the CORDIS dump into a single
consolidated dataset.

Filter logic
------------
  erc_dash_dump   → column "Host Institution(s)"  (substring match, any of DASH_FILTERS)
  erc_cordis_dump → column "org_names"             (substring match, any of CORDIS_FILTERS)

Merge logic
-----------
  Matched on: erc_dash  "Project Number"  ↔  erc_cordis  "id"
  • Rows present in both datasets  → merged single row  (source = "both")
  • Rows only in the DASH dataset  → dash columns only   (source = "dash_only")
  • Rows only in the CORDIS dataset→ cordis columns only (source = "cordis_only")

Column mapping
--------------
  Equivalent columns from both sources are coalesced into a single column.
  Remaining unique columns are renamed to clean, unified names.
  Internal CORDIS metadata (language, modelVersion, …) is dropped.

Output
------
  data/consolidated_erc_<YYYYMMDD_HHMMSS>.xlsx
"""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table as XlTable, TableStyleInfo
from rich.console import Console
from rich.table import Table

console = Console()

# ---------------------------------------------------------------------------
# Configuration – edit these lists to adjust the institution filters
# ---------------------------------------------------------------------------

DASH_FILTERS: list[str] = [
    "Heidelberg University Hospital [999841081,DE]",
    "University of Heidelberg [999987648,DE]",
    "University of Mannheim [999878135,DE]",
]

CORDIS_FILTERS: list[str] = [
    "UNIVERSITAET MANNHEIM",
    "UNIVERSITATSKLINIKUM HEIDELBERG",
    "RUPRECHT-KARLS-UNIVERSITAET HEIDELBERG",
    "ZENTRALINSTITUT FUER SEELISCHE GESUNDHEIT",
]

# Input files
DATA_DIR    = Path("data")
DASH_FILE   = DATA_DIR / "erc_dash_dump_260626.xlsx"
CORDIS_FILE = DATA_DIR / "erc_cordis_dump_260626.xlsx"

# Common join-key name used internally during the merge
_JOIN_KEY = "project_id"

# ---------------------------------------------------------------------------
# Column mapping
# ---------------------------------------------------------------------------
# Columns that appear in both sources with identical/equivalent data.
# Each entry: unified_name -> (preferred_col, fallback_col)
# The first non-null value wins (coalesce).
#
# Prefixes used internally after the merge step:
#   dash_<original_dash_col>   e.g. dash_Acronym
#   cordis_<original_cordis_col>  e.g. cordis_acronym
#
COALESCE_MAP: dict[str, tuple[str, str]] = {
    # unified name        preferred (first non-null)   fallback
    "acronym":            ("dash_Acronym",              "cordis_acronym"),
    "title":              ("cordis_title",              "dash_Project Title"),
    "objective":          ("cordis_objective",          "dash_Abstract"),
    "start_date":         ("cordis_startDate",          "dash_Start Date"),
    "end_date":           ("cordis_endDate",            "dash_End Date"),
    "eu_contribution":    ("cordis_ecMaxContribution",  "dash_EU contribution"),
    "framework_programme":("cordis_frameworkProgramme", "dash_Programme"),
    "call":               ("dash_Call",                 "cordis_call_identifier"),
}

# Single-source columns kept in the output, renamed to clean names.
# Format: old_prefixed_name -> new_clean_name
RENAME_MAP: dict[str, str] = {
    # ── from DASH ──────────────────────────────────────────────────────────
    "dash_Researcher(s)":     "principal_investigator",
    "dash_Host Institution(s)": "host_institutions",
    "dash_Grant Type":        "grant_type",
    "dash_Call Year":         "call_year",
    "dash_Country":           "country_name",
    "dash_Region":            "region",
    "dash_Domain":            "domain",
    "dash_Panel":             "panel",
    "dash_CORDIS Link":       "cordis_link",
    # ── from CORDIS ────────────────────────────────────────────────────────
    "cordis_rcn":                        "rcn",
    "cordis_grantDoi":                   "grant_doi",
    "cordis_status":                     "status",
    "cordis_duration":                   "duration",
    "cordis_totalCost":                  "total_cost",
    "cordis_coordinator_name":           "coordinator_name",
    "cordis_coordinator_shortName":      "coordinator_short_name",
    "cordis_coordinator_country":        "country_code",
    "cordis_coordinator_ecContribution": "coordinator_ec_contribution",
    "cordis_coordinator_activityType":   "coordinator_activity_type",
    "cordis_coordinator_nutsCode":       "nuts_code",
    "cordis_coordinator_nutsName":       "nuts_name",
    "cordis_org_names":                  "org_names",
    "cordis_org_countries":              "org_countries",
    "cordis_programme_code":             "programme_code",
    "cordis_topic_code":                 "topic_code",
    "cordis_topic_title":                "topic_title",
    "cordis_call_title":                 "call_title",
    "cordis_keywords":                   "keywords",
    "cordis_ecSignatureDate":            "ec_signature_date",
    "cordis_terminationDate":            "termination_date",
}

# Columns consumed by COALESCE_MAP or redundant → removed from output
_COALESCE_INPUTS: set[str] = {col for pair in COALESCE_MAP.values() for col in pair}

# Internal CORDIS metadata not relevant for analysis → dropped
_DROP_CORDIS_META: set[str] = {
    "cordis_contenttype",
    "cordis_language",
    "cordis_availableLanguages",
    "cordis_teaser",           # identical to cordis_objective
    "cordis_programme_title",  # too verbose; programme_code is sufficient
    "cordis_modelVersion",
    "cordis_icaVersion",
    "cordis_contentCreationDate",
    "cordis_contentUpdateDate",
    "cordis_lastUpdateDate",
    "cordis_archivedDate",
    "cordis_sourceUpdateDate",
    "cordis_id",               # = project_id
    "dash_Project Number",     # = project_id
}

# Grant type keyword patterns derived from topic_title (used to fill missing grant_type)
# Evaluated in order – first match wins.
GRANT_TYPE_PATTERNS: list[tuple[str, str]] = [
    ("Proof of Concept",  "Proof of Concept"),
    ("Synergy",           "Synergy Grants"),
    ("Consolidator",      "Consolidator Grants"),
    ("Advanced",          "Advanced Grants"),
    ("Starting",          "Starting Grants"),
]

# Desired column order in the final output.
# The first 10 columns are pinned as requested; the remainder follow in logical groups.
_COLUMN_ORDER: list[str] = [
    # ── Pinned first columns ──────────────────────────────────────────────
    "project_id",
    "acronym",
    "principal_investigator",
    "title",
    "grant_type",
    "start_date",
    "end_date",
    "call_year",
    "call",
    "call_title",
    "domain",
    "panel",
    # ── PI & institutions ─────────────────────────────────────────────────
    "host_institutions",
    "org_names",
    "org_countries",
    "coordinator_name",
    "coordinator_short_name",
    # ── Core project ──────────────────────────────────────────────────────
    "objective",
    "status",
    "duration",
    "eu_contribution",
    "total_cost",
    # ── Programme & call ──────────────────────────────────────────────────
    "framework_programme",
    "programme_code",
    "topic_code",
    "topic_title",
    "keywords",
    # -- country
    "country_name",
    "country_code",
    "region",
    "nuts_code",
    "nuts_name",
    "coordinator_ec_contribution",
    "coordinator_activity_type",
    # ── Dates & tracking ──────────────────────────────────────────────────
    "ec_signature_date",
    "termination_date",
        # ── Identifiers ───────────────────────────────────────────────────────
    "source",
    "rcn",
    "grant_doi",
    "cordis_link",
]


# ---------------------------------------------------------------------------
# Helpers – load & filter
# ---------------------------------------------------------------------------

def _substring_filter(df: pd.DataFrame, column: str, substrings: list[str]) -> pd.DataFrame:
    """Return rows where *column* contains at least one string from *substrings*."""
    mask = df[column].apply(
        lambda val: any(s in str(val) for s in substrings) if pd.notna(val) else False
    )
    return df[mask].copy()


def _load_and_filter_dash(path: Path) -> pd.DataFrame:
    console.print(f"[cyan]Loading ERC Dashboard dump:[/] {path}")
    df = pd.read_excel(path, engine="openpyxl")
    console.print(f"  → {len(df):,} rows total")

    filtered = _substring_filter(df, "Host Institution(s)", DASH_FILTERS)
    console.print(f"  → {len(filtered):,} rows after filtering on Host Institution(s)")

    filtered[_JOIN_KEY] = pd.to_numeric(filtered["Project Number"], errors="coerce")
    rename_map = {c: f"dash_{c}" for c in filtered.columns if c != _JOIN_KEY}
    filtered.rename(columns=rename_map, inplace=True)
    return filtered


def _load_and_filter_cordis(path: Path) -> pd.DataFrame:
    console.print(f"[cyan]Loading ERC CORDIS dump:[/]  {path}")
    df = pd.read_excel(path, engine="openpyxl")
    console.print(f"  → {len(df):,} rows total")

    filtered = _substring_filter(df, "org_names", CORDIS_FILTERS)
    console.print(f"  → {len(filtered):,} rows after filtering on org_names")

    filtered[_JOIN_KEY] = pd.to_numeric(filtered["id"], errors="coerce")
    rename_map = {c: f"cordis_{c}" for c in filtered.columns if c != _JOIN_KEY}
    filtered.rename(columns=rename_map, inplace=True)
    return filtered


# ---------------------------------------------------------------------------
# Merge
# ---------------------------------------------------------------------------

def _merge(dash: pd.DataFrame, cordis: pd.DataFrame) -> pd.DataFrame:
    """Full outer merge on *project_id*; adds a 'source' indicator column."""
    merged = pd.merge(
        dash,
        cordis,
        on=_JOIN_KEY,
        how="outer",
        indicator=True,
    )
    source_map = {"both": "both", "left_only": "dash_only", "right_only": "cordis_only"}
    merged["source"] = merged["_merge"].map(source_map)
    merged.drop(columns=["_merge"], inplace=True)
    return merged


# ---------------------------------------------------------------------------
# Reshape – coalesce equivalent columns, rename, reorder
# ---------------------------------------------------------------------------

def _reshape(df: pd.DataFrame) -> pd.DataFrame:
    """
    Coalesce duplicate columns, rename single-source columns, drop metadata.
    Returns a lean, logically-ordered DataFrame.
    """
    out = df.copy()

    # 1. Coalesce equivalent column pairs into unified columns
    for unified, (preferred, fallback) in COALESCE_MAP.items():
        pref_col = out[preferred] if preferred in out.columns else pd.Series(pd.NA, index=out.index)
        fall_col = out[fallback]  if fallback  in out.columns else pd.Series(pd.NA, index=out.index)
        out[unified] = pref_col.where(pref_col.notna(), fall_col)

    # 2. Drop coalesce source columns + metadata
    cols_to_drop = (_COALESCE_INPUTS | _DROP_CORDIS_META) & set(out.columns)
    out.drop(columns=list(cols_to_drop), inplace=True)

    # 3. Rename single-source columns to clean names
    rename = {old: new for old, new in RENAME_MAP.items() if old in out.columns}
    out.rename(columns=rename, inplace=True)

    # 4. Reorder columns: put known columns first in defined order, append any extras
    ordered = [c for c in _COLUMN_ORDER if c in out.columns]
    extras  = [c for c in out.columns  if c not in _COLUMN_ORDER]
    if extras:
        console.print(f"[dim]Reshape: {len(extras)} unmapped column(s) appended: {extras}[/]")
    out = out[ordered + extras]

    # 5. Post-processing: fill derived fields
    #    a) Infer grant_type from topic_title where missing
    if "grant_type" in out.columns and "topic_title" in out.columns:
        def _infer_grant_type(row: pd.Series) -> object:
            if pd.notna(row["grant_type"]):
                return row["grant_type"]
            if pd.isna(row["topic_title"]):
                return pd.NA
            tt = str(row["topic_title"])
            for pattern, label in GRANT_TYPE_PATTERNS:
                if pattern.lower() in tt.lower():
                    return label
            return pd.NA

        out["grant_type"] = out.apply(_infer_grant_type, axis=1)

    #    b) Construct cordis_link from project_id where missing
    if "cordis_link" in out.columns and "project_id" in out.columns:
        base_url = "https://cordis.europa.eu/project/id/"
        constructed = out["project_id"].apply(
            lambda pid: f"{base_url}{int(pid)}" if pd.notna(pid) else pd.NA
        )
        out["cordis_link"] = out["cordis_link"].where(out["cordis_link"].notna(), constructed)

    # 6. Convert start_date / end_date to timezone-naive Python datetimes
    #    so openpyxl writes them as native Excel date serial numbers.
    for date_col in ("start_date", "end_date"):
        if date_col in out.columns:
            out[date_col] = (
                pd.to_datetime(out[date_col], errors="coerce")
                  .dt.tz_localize(None)          # strip tz if present
                  .dt.to_pydatetime()            # → numpy array of datetime | NaT
            )
            # Replace NaT with None so openpyxl leaves the cell empty
            out[date_col] = out[date_col].where(out[date_col].notna(), other=None)

    return out


# ---------------------------------------------------------------------------
# Summary & save
# ---------------------------------------------------------------------------


def _print_summary(merged: pd.DataFrame) -> None:
    counts = merged["source"].value_counts()

    table = Table(title="Consolidation Summary", show_header=True, header_style="bold magenta")
    table.add_column("Source",      style="cyan",  no_wrap=True)
    table.add_column("Rows",        style="green", justify="right")
    table.add_column("Description")

    descriptions = {
        "both":        "Matched in DASH + CORDIS",
        "dash_only":   "Only in ERC Dashboard (no CORDIS record)",
        "cordis_only": "Only in CORDIS (no Dashboard record)",
    }
    for src in ["both", "dash_only", "cordis_only"]:
        n = counts.get(src, 0)
        table.add_row(src, str(n), descriptions[src])
    table.add_row("[bold]TOTAL[/]", f"[bold]{len(merged)}[/]", "", end_section=True)

    console.print()
    console.print(table)
    console.print(f"  Output columns: {merged.shape[1]}")


def _save(merged: pd.DataFrame, out_dir: Path) -> Path:
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = out_dir / f"consolidated_erc_{ts}.xlsx"
    out_dir.mkdir(parents=True, exist_ok=True)

    # Write data with pandas
    merged.to_excel(path, index=False, engine="openpyxl", sheet_name="ERC Data")

    # Reopen with openpyxl to add a native Excel table
    wb = load_workbook(path)
    ws = wb.active

    n_rows = len(merged) + 1                  # +1 for the header row
    last_col = get_column_letter(len(merged.columns))
    table_ref = f"A1:{last_col}{n_rows}"

    tbl = XlTable(displayName="ERCData", ref=table_ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tbl)
    wb.save(path)

    console.print(f"\n[bold green]✓ Saved {len(merged)} rows → {path}[/]")
    return path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    console.rule("[bold blue]ERC Data Consolidation[/]")

    for f in (DASH_FILE, CORDIS_FILE):
        if not f.exists():
            console.print(f"[bold red]ERROR:[/] File not found: {f}")
            sys.exit(1)

    dash   = _load_and_filter_dash(DASH_FILE)
    cordis = _load_and_filter_cordis(CORDIS_FILE)

    console.print("\n[cyan]Merging datasets…[/]")
    merged = _merge(dash, cordis)

    console.print("[cyan]Reshaping columns…[/]")
    merged = _reshape(merged)

    _print_summary(merged)
    _save(merged, DATA_DIR)


if __name__ == "__main__":
    main()
