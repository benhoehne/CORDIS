#!/usr/bin/env python3
"""
CORDIS Extraction Runner
=========================
Reads ``config.json`` and runs a CORDIS data extraction.

Workflow
--------
1. Paste your finalised CORDIS query string into the ``"query"`` field of
   ``config.json`` (copy it directly from the CORDIS search platform).
2. Adjust ``"summary_fields"`` in ``config.json`` to control which columns
   appear in the auto-generated summary spreadsheet.
3. Run:

       python run.py

Options
-------
    python run.py --show-query              # print query, do not submit
    python run.py --archived                # include archived records
    python run.py --save-format csv         # full export as CSV instead of XLSX
    python run.py --config other.json       # use a different config file
    python run.py --output data/my_output   # custom output base path
"""

import sys
import json
from pathlib import Path
from datetime import datetime

import click
from dotenv import load_dotenv
from rich.console import Console
from rich.panel import Panel
from rich.syntax import Syntax

sys.path.insert(0, str(Path(__file__).parent))

from src.cordis_client import CordisClient, CordisAPIError
from src.query_builder import load_config
from src.data_processor import (
    parse_response,
    print_summary,
    basic_analytics,
    save_dataframe,
    save_raw,
)

load_dotenv()
console = Console()


# ── CLI ───────────────────────────────────────────────────────────────────────

@click.command()
@click.option(
    "--key",
    envvar="CORDIS_API_KEY",
    required=True,
    help="CORDIS API key (or set CORDIS_API_KEY env var).",
)
@click.option(
    "--config", "config_path",
    default=None,
    help="Path to the config JSON file. Defaults to config.json in the project root.",
)
@click.option(
    "--output", "output_path",
    default=None,
    help="Output base path (without extension). Overrides config output_prefix.",
)
@click.option(
    "--save-format", "save_format",
    type=click.Choice(["csv", "xlsx", "json", "parquet"], case_sensitive=False),
    default="xlsx",
    show_default=True,
    help="Format for the full parsed-data export.",
)
@click.option(
    "--archived",
    is_flag=True,
    default=False,
    help="Include archived CORDIS records.",
)
@click.option(
    "--no-save",
    is_flag=True,
    default=False,
    help="Print summary to console only; do not write any files.",
)
@click.option(
    "--show-query",
    is_flag=True,
    default=False,
    help="Print the query string and exit without submitting.",
)
def main(
    key: str,
    config_path: str | None,
    output_path: str | None,
    save_format: str,
    archived: bool,
    no_save: bool,
    show_query: bool,
) -> None:
    """Run a CORDIS data extraction defined in config.json."""

    # ── Load config ───────────────────────────────────────────────────────────
    try:
        cfg = load_config(config_path)
    except (FileNotFoundError, ValueError) as exc:
        console.print(f"[bold red]✗ {exc}[/]")
        sys.exit(1)

    query: str             = cfg["query"]
    summary_fields: list   = cfg.get("summary_fields", [])
    output_prefix: str     = cfg.get("output_prefix", "data/cordis_extraction")

    cfg_file = Path(config_path) if config_path else Path("config.json")

    console.print(
        Panel.fit(
            f"[bold]CORDIS Extraction Runner[/]\n\n"
            f"[dim]Config:[/]    {cfg_file.name}\n"
            f"[dim]Save as:[/]   {save_format.upper()}\n"
            f"[dim]Archived:[/]  {archived}",
            border_style="cyan",
        )
    )

    if show_query:
        console.print("\n[bold yellow]Query:[/]")
        console.print(Syntax(query, "text", theme="monokai", word_wrap=True))
        return

    console.print("\n[bold yellow]Query:[/]")
    console.print(f"[dim]{query}[/]\n")

    # ── Submit & download ─────────────────────────────────────────────────────
    try:
        client    = CordisClient(api_key=key)
        raw_bytes = client.extract(query=query, output_format="json", archived=archived)
    except CordisAPIError as exc:
        console.print(f"[bold red]✗ CORDIS API Error:[/] {exc}")
        sys.exit(1)
    except TimeoutError as exc:
        console.print(f"[bold red]✗ Timeout:[/] {exc}")
        sys.exit(1)
    except Exception as exc:
        console.print(f"[bold red]✗ Unexpected error:[/] {exc}")
        raise

    # ── Parse ─────────────────────────────────────────────────────────────────
    df = parse_response(raw_bytes, "json")
    print_summary(df)

    stats = basic_analytics(df)
    _print_analytics(stats)

    if no_save:
        console.print("\n[dim]--no-save: skipping file output.[/]")
        return

    # ── Save ──────────────────────────────────────────────────────────────────
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = output_path or f"{output_prefix}_{timestamp}"

    save_dataframe(df, base, fmt=save_format)
    save_raw(raw_bytes, Path(base).with_suffix(".zip"))

    if summary_fields:
        _save_summary_xlsx(df, base, summary_fields)
    else:
        console.print("[dim]No summary_fields in config; skipping summary export.[/]")

    console.print(f"\n[bold green]Done![/] Saved to [underline]{base}.{save_format}[/]")


# ── Helpers ───────────────────────────────────────────────────────────────────

def _save_summary_xlsx(df, base_path: str, fields: list[str]) -> None:
    """
    Write a summary XLSX with:
    • Only the columns listed in *fields* (absent ones are skipped)
    • A ``Link`` column with a clickable CORDIS project URL per row
    • The sheet wrapped in a native Excel Table (filterable / sortable)
    """
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    # Select columns
    present  = [f for f in fields if f in df.columns]
    missing  = [f for f in fields if f not in df.columns]
    if missing:
        console.print(
            f"[dim]Summary: skipping {len(missing)} absent field(s): "
            f"{', '.join(missing)}[/]"
        )

    summary_df = df[present].copy()

    # Add Link column
    id_col = df.get("id") if "id" not in summary_df.columns else summary_df["id"]
    if id_col is not None:
        summary_df["Link"] = id_col.apply(
            lambda x: (
                f"https://cordis.europa.eu/project/id/{x}"
                if pd.notna(x) and str(x).strip()
                else ""
            )
        )

    out = Path(base_path).with_name(Path(base_path).stem + "_summary.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    summary_df.to_excel(out, index=False, engine="openpyxl")

    # Apply Excel Table + hyperlinks
    wb = load_workbook(out)
    ws = wb.active
    n_rows, n_cols = len(summary_df), len(summary_df.columns)
    last_col = get_column_letter(n_cols)

    tbl = Table(displayName="CORDISProjects", ref=f"A1:{last_col}{n_rows + 1}")
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tbl)

    if "Link" in summary_df.columns:
        link_col_letter = get_column_letter(summary_df.columns.get_loc("Link") + 1)
        for row_idx, url in enumerate(summary_df["Link"], start=2):
            if url:
                cell = ws[f"{link_col_letter}{row_idx}"]
                cell.hyperlink = url
                cell.value     = url
                cell.style     = "Hyperlink"

    wb.save(out)
    console.print(
        f"[bold green]✓ Saved summary table "
        f"({len(summary_df.columns)} cols, {n_rows} rows) → {out}[/]"
    )


def _print_analytics(stats: dict) -> None:
    """Pretty-print analytics."""
    console.rule("[bold blue]Analytics[/]")
    console.print(f"[green]Total records:[/] {stats.get('total_projects', 'N/A')}")

    for key in ("ecMaxContribution", "totalCost"):
        if f"total_{key}" in stats:
            console.print(
                f"[green]{key}:[/] "
                f"total = [bold]€{stats[f'total_{key}']:,.0f}[/], "
                f"mean = [bold]€{stats[f'mean_{key}']:,.0f}[/]"
            )

    for col in ("status", "fundingScheme", "projectStatus"):
        by_key = f"by_{col}"
        if by_key in stats:
            console.print(f"\n[bold]By {col}:[/]")
            for val, count in sorted(stats[by_key].items(), key=lambda x: -x[1]):
                console.print(f"  {val:<30} {count:>4}")


if __name__ == "__main__":
    main()
